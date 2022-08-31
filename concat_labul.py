# import os
import pandas as pd
from openpyxl.workbook import workbook
from openpyxl import load_workbook
from tqdm import tqdm

neraca = []

for x in tqdm(range(1,922)):
    wb = load_workbook(f'd:/ojk/labul juni 2022/labul_bpr_{x}.xlsx')
    ws = wb.active
    try:
        kabupaten = ws['b10'].value.split(',')[1].strip()
    except:
        kabupaten = ' '
    data = {
        'Nama BPR'              : ws['d7'].value[9:],
        'Kode Bank'             : ws['d7'].value[0:7],
        'Provinsi'              : ws['b10'].value.split(',')[0],
        'Kabupaten/Kota'        : kabupaten,
        'Total Aset'            : ws['g23'].value,
        'Tabungan'              : ws['g44'].value,
        'Deposito'              : ws['g45'].value,
        'Kredit yg diberikan'   : ws['g27'].value + ws['g28'].value - ws['g29'].value,
        'Modal Disetor'         : ws['g53'].value - ws['g54'].value,
        'Cadangan'              : ws['g66'].value + ws['g67'].value,
        'Laba Tahun lalu'       : ws['g69'].value,
        '50%_Laba tahun berjalan': (ws['g70'].value)/2,
        'ABP Simpanan'          : ws['g46'].value,
        'ABP Pinjaman'          : ws['g47'].value 
    }
    neraca.append(data)

df = pd.DataFrame(neraca)
print(df)
df.to_csv('d:/ojk/neraca_BPR_Juni_2022.csv')