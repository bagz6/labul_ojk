# import os
import pandas as pd
from openpyxl.workbook import workbook
from openpyxl import load_workbook
from tqdm import tqdm

neraca = []

for x in tqdm(range(1,918)):
    wb = load_workbook(f'd:/ojk/labul kap juni 2022/labul_kap_bpr_{x}.xlsx')
    ws = wb.active

    data = {
        'Nama BPR'              : ws['c7'].value[9:],
        'Kode Bank'             : ws['c7'].value[0:7],
        'KPMM'                  : ws['O28'].value,
        'KAP'                   : ws['o29'].value,
        'PPAP'                  : ws['o30'].value,
        'ROA'                   : ws['o32'].value,
        'NPL'                   : ws['o31'].value,
        'BOPO'                  : ws['o33'].value,
        'LDR'                   : ws['o34'].value,
        'CR'                    : ws['o35'].value
    }
    neraca.append(data)

df = pd.DataFrame(neraca)
print(df)
df.to_csv('d:/ojk/kap_BPR_Juni_2022.csv')